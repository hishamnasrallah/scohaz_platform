import csv
import json
import io
from datetime import datetime, date
from decimal import Decimal
from typing import List, Dict, Any, Union

from django.http import HttpResponse
from django.utils import timezone
from django.template.loader import render_to_string
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from reporting.models import Report, ReportField


class ReportExporter:
    """
    Handles exporting report data to various formats.
    Supports CSV, Excel, PDF, and HTML formats.
    """

    def __init__(self):
        self.styles = getSampleStyleSheet()
        self._setup_pdf_styles()

    def _setup_pdf_styles(self):
        """Setup custom PDF styles."""
        # Title style
        self.title_style = ParagraphStyle(
            'CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1a1a1a'),
            spaceAfter=30,
            alignment=TA_CENTER
        )

        # Subtitle style
        self.subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=self.styles['Normal'],
            fontSize=12,
            textColor=colors.HexColor('#666666'),
            spaceAfter=20,
            alignment=TA_CENTER
        )

        # Header style for tables
        self.header_style = ParagraphStyle(
            'CustomHeader',
            parent=self.styles['Normal'],
            fontSize=10,
            textColor=colors.white,
            alignment=TA_CENTER
        )

    def export_csv(self, report: Report, data: List[Dict[str, Any]],
                   include_headers: bool = True,
                   delimiter: str = ',') -> HttpResponse:
        """
        Export report data to CSV format.
        
        Args:
            report: Report instance
            data: List of data rows
            include_headers: Whether to include column headers
            delimiter: CSV delimiter character
            
        Returns:
            HttpResponse with CSV content
        """
        output = io.StringIO()

        # Get visible fields
        fields = report.fields.filter(is_visible=True).order_by('order')

        # Create CSV writer
        writer = csv.writer(output, delimiter=delimiter, quoting=csv.QUOTE_MINIMAL)

        # Write headers
        if include_headers:
            headers = [field.display_name for field in fields]
            writer.writerow(headers)

        # Write data rows
        for row in data:
            csv_row = []
            for field in fields:
                value = self._get_field_value(row, field)
                csv_row.append(self._format_csv_value(value))
            writer.writerow(csv_row)

        # Create response
        response = HttpResponse(output.getvalue(), content_type='text/csv')
        filename = self._sanitize_filename(report.name)
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'

        return response

    def export_excel(self, report: Report, data: List[Dict[str, Any]],
                     include_formatting: bool = True,
                     include_summary: bool = True) -> HttpResponse:
        """
        Export report data to Excel format with formatting.
        
        Args:
            report: Report instance
            data: List of data rows
            include_formatting: Whether to apply cell formatting
            include_summary: Whether to include summary sheet
            
        Returns:
            HttpResponse with Excel content
        """
        # Create workbook
        wb = openpyxl.Workbook()

        # Main data sheet
        ws = wb.active
        ws.title = self._sanitize_sheet_name(report.name)

        # Get visible fields
        fields = report.fields.filter(is_visible=True).order_by('order')

        # Write title
        ws.merge_cells('A1:' + get_column_letter(len(fields)) + '1')
        title_cell = ws['A1']
        title_cell.value = report.name
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Write metadata
        ws['A2'] = f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A3'] = f"Total Records: {len(data)}"

        # Headers row
        header_row = 5
        for col_idx, field in enumerate(fields, 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.value = field.display_name

            if include_formatting:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')

                # Set column width based on field configuration
                if field.width:
                    ws.column_dimensions[get_column_letter(col_idx)].width = field.width / 7
                else:
                    ws.column_dimensions[get_column_letter(col_idx)].width = 15

        # Data rows
        for row_idx, row_data in enumerate(data, header_row + 1):
            for col_idx, field in enumerate(fields, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                value = self._get_field_value(row_data, field)
                cell.value = self._format_excel_value(value)

                if include_formatting:
                    # Apply formatting based on field type
                    if field.field_type in ['IntegerField', 'BigIntegerField']:
                        cell.number_format = '#,##0'
                    elif field.field_type in ['DecimalField', 'FloatField']:
                        decimals = field.formatting.get('decimals', 2) if field.formatting else 2
                        cell.number_format = f'#,##0.{"0" * decimals}'
                    elif field.field_type in ['DateField']:
                        cell.number_format = 'YYYY-MM-DD'
                    elif field.field_type in ['DateTimeField']:
                        cell.number_format = 'YYYY-MM-DD HH:MM:SS'

                    # Apply custom formatting
                    if field.formatting:
                        format_type = field.formatting.get('type')
                        if format_type == 'currency':
                            prefix = field.formatting.get('prefix', '$')
                            decimals = field.formatting.get('decimals', 2)
                            cell.number_format = f'{prefix}#,##0.{"0" * decimals}'
                        elif format_type == 'percentage':
                            decimals = field.formatting.get('decimals', 1)
                            cell.number_format = f'0.{"0" * decimals}%'

                    # Alignment based on field type
                    if field.field_type in ['IntegerField', 'DecimalField', 'FloatField']:
                        cell.alignment = Alignment(horizontal='right')
                    elif field.field_type in ['BooleanField']:
                        cell.alignment = Alignment(horizontal='center')

        # Add borders if formatting enabled
        if include_formatting:
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for row in ws.iter_rows(min_row=header_row, max_row=ws.max_row,
                                    min_col=1, max_col=len(fields)):
                for cell in row:
                    cell.border = border

        # Add autofilter
        ws.auto_filter.ref = ws.dimensions

        # Freeze panes (header row)
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

        # Add summary sheet if requested
        if include_summary and len(data) > 0:
            self._add_summary_sheet(wb, report, data, fields)

        # Save to response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = self._sanitize_filename(report.name)
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'

        wb.save(response)
        return response

    def export_pdf(self, report: Report, data: List[Dict[str, Any]],
                   page_size: str = 'letter',
                   orientation: str = 'portrait') -> HttpResponse:
        """
        Export report data to PDF format.
        
        Args:
            report: Report instance
            data: List of data rows
            page_size: Page size ('letter' or 'a4')
            orientation: Page orientation ('portrait' or 'landscape')
            
        Returns:
            HttpResponse with PDF content
        """
        # Create response
        response = HttpResponse(content_type='application/pdf')
        filename = self._sanitize_filename(report.name)
        response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'

        # Set page size
        if page_size == 'a4':
            pagesize = A4
        else:
            pagesize = letter

        if orientation == 'landscape':
            pagesize = landscape(pagesize)

        # Create PDF document
        doc = SimpleDocTemplate(
            response,
            pagesize=pagesize,
            rightMargin=0.5*inch,
            leftMargin=0.5*inch,
            topMargin=0.5*inch,
            bottomMargin=0.5*inch
        )

        # Container for elements
        elements = []

        # Add title
        title = Paragraph(report.name, self.title_style)
        elements.append(title)

        # Add subtitle with metadata
        subtitle_text = f"Generated on {timezone.now().strftime('%B %d, %Y at %I:%M %p')}<br/>"
        subtitle_text += f"Total Records: {len(data)}"
        subtitle = Paragraph(subtitle_text, self.subtitle_style)
        elements.append(subtitle)
        elements.append(Spacer(1, 0.3*inch))

        # Get visible fields
        fields = report.fields.filter(is_visible=True).order_by('order')

        # Prepare table data
        table_data = []

        # Headers
        headers = []
        for field in fields:
            header_text = Paragraph(field.display_name, self.header_style)
            headers.append(header_text)
        table_data.append(headers)

        # Data rows
        for row_data in data:
            row = []
            for field in fields:
                value = self._get_field_value(row_data, field)
                formatted_value = self._format_pdf_value(value, field)

                # Create paragraph for cell
                cell_style = self.styles['Normal']
                cell_style.fontSize = 8

                # Alignment based on field type
                if field.field_type in ['IntegerField', 'DecimalField', 'FloatField']:
                    cell_style.alignment = TA_RIGHT

                cell = Paragraph(str(formatted_value), cell_style)
                row.append(cell)
            table_data.append(row)

        # Create table
        col_widths = self._calculate_column_widths(fields, pagesize[0] - inch)
        table = Table(table_data, colWidths=col_widths, repeatRows=1)

        # Apply table style
        table_style = TableStyle([
            # Header style
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),

            # Data rows style
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ])

        table.setStyle(table_style)
        elements.append(table)

        # Build PDF
        doc.build(elements)
        return response

    def export_html(self, report: Report, data: List[Dict[str, Any]],
                    template_name: str = 'reporting/export/report.html',
                    include_styles: bool = True) -> HttpResponse:
        """
        Export report data to HTML format.
        
        Args:
            report: Report instance
            data: List of data rows
            template_name: Template to use for rendering
            include_styles: Whether to include inline CSS styles
            
        Returns:
            HttpResponse with HTML content
        """
        # Get visible fields
        fields = report.fields.filter(is_visible=True).order_by('order')

        # Prepare formatted data
        formatted_data = []
        for row_data in data:
            formatted_row = {}
            for field in fields:
                value = self._get_field_value(row_data, field)
                formatted_row[field.field_path] = {
                    'value': value,
                    'display': self._format_html_value(value, field),
                    'field': field
                }
            formatted_data.append(formatted_row)

        # Render template
        context = {
            'report': report,
            'fields': fields,
            'data': formatted_data,
            'generated_at': timezone.now(),
            'total_records': len(data),
            'include_styles': include_styles,
        }

        html_content = render_to_string(template_name, context)

        response = HttpResponse(html_content, content_type='text/html')
        return response

    def export_json(self, report: Report, data: List[Dict[str, Any]],
                    pretty_print: bool = True,
                    include_metadata: bool = True) -> HttpResponse:
        """
        Export report data to JSON format.
        
        Args:
            report: Report instance
            data: List of data rows
            pretty_print: Whether to format JSON with indentation
            include_metadata: Whether to include report metadata
            
        Returns:
            HttpResponse with JSON content
        """
        # Prepare export data
        export_data = {
            'data': data,
        }

        if include_metadata:
            fields = report.fields.filter(is_visible=True).order_by('order')
            export_data['metadata'] = {
                'report_name': report.name,
                'report_description': report.description,
                'generated_at': timezone.now().isoformat(),
                'total_records': len(data),
                'fields': [
                    {
                        'name': field.field_path,
                        'display_name': field.display_name,
                        'type': field.field_type,
                        'aggregation': field.aggregation,
                    }
                    for field in fields
                ],
            }

        # Serialize to JSON
        indent = 2 if pretty_print else None
        json_content = json.dumps(export_data, indent=indent, default=str)

        response = HttpResponse(json_content, content_type='application/json')
        filename = self._sanitize_filename(report.name)
        response['Content-Disposition'] = f'attachment; filename="{filename}.json"'

        return response

    # Helper methods

    def _get_field_value(self, row_data: Dict[str, Any], field: ReportField) -> Any:
        """Extract field value from row data."""
        if isinstance(row_data, dict):
            # Try direct field path
            if field.field_path in row_data:
                return row_data[field.field_path]

            # Try display name
            if field.display_name in row_data:
                return row_data[field.display_name]

            # Try navigating through nested structure
            parts = field.field_path.split('__')
            value = row_data
            for part in parts:
                if isinstance(value, dict) and part in value:
                    value = value[part]
                else:
                    return None
            return value

        return None

    def _format_csv_value(self, value: Any) -> str:
        """Format value for CSV export."""
        if value is None:
            return ''
        elif isinstance(value, bool):
            return 'Yes' if value else 'No'
        elif isinstance(value, (datetime, date)):
            return value.isoformat()
        elif isinstance(value, Decimal):
            return str(value)
        else:
            return str(value)

    def _format_excel_value(self, value: Any) -> Any:
        """Format value for Excel export."""
        if value is None:
            return ''
        elif isinstance(value, bool):
            return value
        elif isinstance(value, datetime):
            return value.replace(tzinfo=None) if value.tzinfo else value
        elif isinstance(value, date):
            return value
        elif isinstance(value, Decimal):
            return float(value)
        else:
            return value

    def _format_pdf_value(self, value: Any, field: ReportField) -> str:
        """Format value for PDF export."""
        if value is None:
            return ''

        # Apply field formatting
        if field.formatting:
            format_type = field.formatting.get('type')

            if format_type == 'currency' and isinstance(value, (int, float, Decimal)):
                prefix = field.formatting.get('prefix', '$')
                decimals = field.formatting.get('decimals', 2)
                return f"{prefix}{float(value):,.{decimals}f}"

            elif format_type == 'percentage' and isinstance(value, (int, float, Decimal)):
                decimals = field.formatting.get('decimals', 1)
                return f"{float(value):.{decimals}f}%"

            elif format_type == 'number' and isinstance(value, (int, float, Decimal)):
                decimals = field.formatting.get('decimals', 0)
                return f"{float(value):,.{decimals}f}"

        # Default formatting
        if isinstance(value, bool):
            return 'Yes' if value else 'No'
        elif isinstance(value, datetime):
            return value.strftime('%Y-%m-%d %H:%M')
        elif isinstance(value, date):
            return value.strftime('%Y-%m-%d')
        elif isinstance(value, Decimal):
            return f"{float(value):,.2f}"
        else:
            return str(value)

    def _format_html_value(self, value: Any, field: ReportField) -> str:
        """Format value for HTML export."""
        # Similar to PDF formatting but with HTML escaping
        from django.utils.html import escape
        formatted = self._format_pdf_value(value, field)
        return escape(formatted)

    def _sanitize_filename(self, filename: str) -> str:
        """Sanitize filename for file system."""
        import re
        # Remove invalid characters
        filename = re.sub(r'[^\w\s-]', '', filename)
        # Replace spaces with underscores
        filename = re.sub(r'[-\s]+', '_', filename)
        # Limit length
        return filename[:50]

    def _sanitize_sheet_name(self, name: str) -> str:
        """Sanitize sheet name for Excel."""
        # Excel sheet names have restrictions
        invalid_chars = ['/', '\\', '?', '*', '[', ']', ':']
        for char in invalid_chars:
            name = name.replace(char, '_')
        # Limit to 31 characters
        return name[:31]

    def _calculate_column_widths(self, fields: List[ReportField],
                                 available_width: float) -> List[float]:
        """Calculate column widths for PDF table."""
        # If widths are specified, use them
        specified_widths = [f.width for f in fields if f.width]
        if len(specified_widths) == len(fields):
            # Scale to fit page
            total_specified = sum(specified_widths)
            scale_factor = available_width / total_specified
            return [w * scale_factor for w in specified_widths]

        # Otherwise, distribute evenly
        return [available_width / len(fields)] * len(fields)

    def _add_summary_sheet(self, wb: openpyxl.Workbook, report: Report,
                           data: List[Dict[str, Any]], fields: List[ReportField]):
        """Add a summary sheet to Excel workbook."""
        ws = wb.create_sheet(title="Summary")

        # Title
        ws['A1'] = "Report Summary"
        ws['A1'].font = Font(size=16, bold=True)

        # Basic info
        row = 3
        ws[f'A{row}'] = "Report Name:"
        ws[f'B{row}'] = report.name

        row += 1
        ws[f'A{row}'] = "Generated:"
        ws[f'B{row}'] = timezone.now().strftime('%Y-%m-%d %H:%M:%S')

        row += 1
        ws[f'A{row}'] = "Total Records:"
        ws[f'B{row}'] = len(data)

        # Field summary
        row += 2
        ws[f'A{row}'] = "Field Summary"
        ws[f'A{row}'].font = Font(bold=True)

        row += 1
        ws[f'A{row}'] = "Field Name"
        ws[f'B{row}'] = "Type"
        ws[f'C{row}'] = "Aggregation"

        for cell in [ws[f'A{row}'], ws[f'B{row}'], ws[f'C{row}']]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        for field in fields:
            row += 1
            ws[f'A{row}'] = field.display_name
            ws[f'B{row}'] = field.field_type
            ws[f'C{row}'] = field.aggregation or 'None'

        # Adjust column widths
        for column in ['A', 'B', 'C']:
            ws.column_dimensions[column].width = 20