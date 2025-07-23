import time
from django.contrib.contenttypes.models import ContentType
from django.db.models import Model, Q
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.pagination import PageNumberPagination
from django.http import HttpResponse, JsonResponse
import csv
import json
from openpyxl import Workbook
from django.db import models as django_models

from inquiry.models import InquiryConfiguration, InquiryExecution, InquiryTemplate
from inquiry.services.query_builder import DynamicQueryBuilder
from inquiry.apis.serializers.dynamic import DynamicModelSerializer
from inquiry.apis.serializers.inquiry import (
    InquiryConfigurationSerializer,
    InquiryExecutionSerializer,
    InquiryTemplateSerializer, InquiryPermissionSerializer
)
# Add these imports if not already present
from inquiry.models import (
    InquiryConfiguration, InquiryField, InquiryFilter,
    InquiryRelation, InquirySort, InquiryPermission,
    InquiryExecution, InquiryTemplate
)
from inquiry.apis.serializers.inquiry import (
    InquiryConfigurationSerializer,
    InquiryFieldSerializer,
    InquiryFilterSerializer,
    InquiryRelationSerializer,
    InquirySortSerializer,
    InquiryPermissionSerializer,
    InquiryExecutionSerializer,
    InquiryTemplateSerializer
)

class DynamicPagination(PageNumberPagination):
    page_size_query_param = 'page_size'
    max_page_size = 1000

    def get_page_size(self, request):
        if self.page_size_query_param:
            try:
                inquiry_code = request.parser_context['kwargs'].get('code')
                if inquiry_code:
                    inquiry = InquiryConfiguration.objects.filter(
                        code=inquiry_code
                    ).first()
                    if inquiry:
                        self.page_size = inquiry.default_page_size
                        self.max_page_size = inquiry.max_page_size
            except:
                pass
        return super().get_page_size(request)


class InquiryConfigurationViewSet(viewsets.ModelViewSet):
    """CRUD for inquiry configurations"""
    queryset = InquiryConfiguration.objects.all()
    serializer_class = InquiryConfigurationSerializer
    permission_classes = [IsAuthenticated]
    lookup_field = 'code'

    def get_queryset(self):
        queryset = super().get_queryset()
        if not self.request.user.is_superuser:
            # Filter by user's groups OR public inquiries
            user_groups = self.request.user.groups.all()
            queryset = queryset.filter(
                Q(allowed_groups__in=user_groups) | Q(is_public=True)
            ).distinct()
        return queryset

    @action(detail=True, methods=['get'])
    def preview(self, request, code=None):
        """Preview inquiry structure without executing"""
        inquiry = self.get_object()

        # Get model info
        model = inquiry.content_type.model_class()

        # Build structure
        structure = {
            'model': {
                'app_label': model._meta.app_label,
                'model_name': model._meta.model_name,
                'verbose_name': model._meta.verbose_name,
                'verbose_name_plural': model._meta.verbose_name_plural,
            },
            'fields': [],
            'relations': [],
            'filters': [],
            'sorts': []
        }

        # Add fields
        for field in inquiry.fields.all():
            field_info = {
                'field_path': field.field_path,
                'display_name': field.display_name,
                'field_type': field.field_type,
                'is_sortable': field.is_sortable,
                'is_searchable': field.is_searchable,
                'is_filterable': field.is_filterable,
            }
            structure['fields'].append(field_info)

        # Add relations
        for relation in inquiry.relations.all():
            relation_info = {
                'relation_path': relation.relation_path,
                'display_name': relation.display_name,
                'relation_type': relation.relation_type,
                'include_fields': relation.include_fields,
            }
            structure['relations'].append(relation_info)

        # Add filters
        for filter_obj in inquiry.filters.all():
            filter_info = {
                'code': filter_obj.code,
                'name': filter_obj.name,
                'field_path': filter_obj.field_path,
                'operator': filter_obj.operator,
                'filter_type': filter_obj.filter_type,
            }
            structure['filters'].append(filter_info)

        return Response(structure)

    @action(detail=False, methods=['get'])
    def available_models(self, request):
        """List all available models for inquiry"""
        models = []
        for ct in ContentType.objects.all():
            model_class = ct.model_class()
            if model_class and issubclass(model_class, django_models.Model):
                # Process fields
                fields = []
                for f in model_class._meta.get_fields():
                    field_info = {
                        'name': f.name,
                        'is_relation': f.is_relation
                    }

                    # Handle different field types
                    if hasattr(f, 'get_internal_type'):
                        field_info['type'] = f.get_internal_type()
                    else:
                        # For reverse relations
                        if hasattr(f, 'field'):
                            field_info['type'] = f.field.get_internal_type() + '_reverse'
                        else:
                            field_info['type'] = type(f).__name__

                    # Get verbose name safely
                    if hasattr(f, 'verbose_name'):
                        field_info['verbose_name'] = str(f.verbose_name)
                    elif hasattr(f, 'related_name'):
                        field_info['verbose_name'] = f.related_name or f.name
                    else:
                        field_info['verbose_name'] = f.name.replace('_', ' ').title()

                    # Add relation type for relations
                    if f.is_relation:
                        if hasattr(f, 'many_to_many') and f.many_to_many:
                            field_info['relation_type'] = 'many_to_many'
                        elif hasattr(f, 'one_to_many') and f.one_to_many:
                            field_info['relation_type'] = 'one_to_many'
                        elif hasattr(f, 'one_to_one') and f.one_to_one:
                            field_info['relation_type'] = 'one_to_one'
                        elif hasattr(f, 'many_to_one') and f.many_to_one:
                            field_info['relation_type'] = 'foreign_key'
                        else:
                            field_info['relation_type'] = 'unknown'

                        # Add related model info
                        if hasattr(f, 'related_model') and f.related_model:
                            field_info['related_model'] = f.related_model._meta.label

                    fields.append(field_info)

                models.append({
                    'id': ct.id,
                    'app_label': ct.app_label,
                    'model': ct.model,
                    'name': str(model_class._meta.verbose_name),
                    'verbose_name': str(model_class._meta.verbose_name),
                    'fields': fields
                })

        return Response(models)

    @action(detail=True, methods=['get'])
    def export(self, request, code=None):
        """Export inquiry configuration as JSON"""
        inquiry = self.get_object()

        # Build complete configuration
        export_data = {
            'inquiry': InquiryConfigurationSerializer(inquiry).data,
            'fields': InquiryFieldSerializer(inquiry.fields.all(), many=True).data,
            'filters': InquiryFilterSerializer(inquiry.filters.all(), many=True).data,
            'relations': InquiryRelationSerializer(inquiry.relations.all(), many=True).data,
            'sorts': InquirySortSerializer(inquiry.sorts.all(), many=True).data,
            'permissions': InquiryPermissionSerializer(inquiry.permissions.all(), many=True).data,
        }

        response = JsonResponse(export_data, json_dumps_params={'indent': 2})
        response['Content-Disposition'] = f'attachment; filename="{inquiry.code}_config.json"'
        return response

    @action(detail=False, methods=['post'], url_path='import')
    def import_config(self, request):
        """Import inquiry configuration from JSON file"""
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            # Read and parse JSON
            content = file.read()
            data = json.loads(content)

            # Extract inquiry data
            inquiry_data = data.get('inquiry', {})
            fields_data = data.get('fields', [])
            filters_data = data.get('filters', [])
            relations_data = data.get('relations', [])
            sorts_data = data.get('sorts', [])
            permissions_data = data.get('permissions', [])

            # Remove IDs and timestamps for new creation
            inquiry_data.pop('id', None)
            inquiry_data.pop('created_at', None)
            inquiry_data.pop('updated_at', None)

            # Ensure unique code
            original_code = inquiry_data.get('code', 'imported_inquiry')
            code = original_code
            counter = 1
            while InquiryConfiguration.objects.filter(code=code).exists():
                code = f"{original_code}_{counter}"
                counter += 1
            inquiry_data['code'] = code

            # Create inquiry
            inquiry_data['created_by'] = request.user.id
            inquiry_serializer = InquiryConfigurationSerializer(data=inquiry_data)
            if not inquiry_serializer.is_valid():
                return Response(inquiry_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

            inquiry = inquiry_serializer.save()

            # Create related objects
            # Fields
            for field_data in fields_data:
                field_data.pop('id', None)
                field_data['inquiry'] = inquiry.id
                InquiryField.objects.create(**field_data)

            # Filters
            for filter_data in filters_data:
                filter_data.pop('id', None)
                filter_data['inquiry'] = inquiry.id
                InquiryFilter.objects.create(**filter_data)

            # Relations
            for relation_data in relations_data:
                relation_data.pop('id', None)
                relation_data['inquiry'] = inquiry.id
                InquiryRelation.objects.create(**relation_data)

            # Sorts
            for sort_data in sorts_data:
                sort_data.pop('id', None)
                sort_data['inquiry'] = inquiry.id
                InquirySort.objects.create(**sort_data)

            # Permissions
            for perm_data in permissions_data:
                perm_data.pop('id', None)
                perm_data['inquiry'] = inquiry.id
                InquiryPermission.objects.create(**perm_data)

            return Response(InquiryConfigurationSerializer(inquiry).data, status=status.HTTP_201_CREATED)

        except json.JSONDecodeError:
            return Response({'error': 'Invalid JSON file'}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class InquiryExecutionViewSet(viewsets.ViewSet):
    """Execute configured inquiries"""
    permission_classes = [IsAuthenticated]
    pagination_class = DynamicPagination

    @action(detail=False, methods=['post'], url_path='execute/(?P<code>[^/.]+)')
    def execute(self, request, code=None):
        """Execute an inquiry by code"""
        start_time = time.time()

        try:
            # Get inquiry configuration
            inquiry = InquiryConfiguration.objects.get(code=code, active=True)

            # Check permissions
            if not self.has_permission(request.user, inquiry):
                return Response(
                    {'error': 'Permission denied'},
                    status=status.HTTP_403_FORBIDDEN
                )

            # Get request parameters
            filters = request.data.get('filters', {})
            search = request.data.get('search', '')
            sort = request.data.get('sort', [])
            export_format = request.data.get('export')

            # Build queryset
            builder = DynamicQueryBuilder(inquiry)
            queryset = builder.build_queryset(
                filters=filters,
                search=search,
                sort=sort,
                user=request.user
            )

            # Get aggregations if requested
            include_aggregations = request.data.get('include_aggregations', False)
            aggregations = {}
            if include_aggregations:
                aggregations = builder.get_aggregations(queryset)

            # Handle export
            if export_format and inquiry.allow_export:
                return self.export_data(
                    queryset,
                    inquiry,
                    export_format,
                    request.user
                )

            # Paginate results
            paginator = DynamicPagination()
            page = paginator.paginate_queryset(queryset, request, view=self)

            # Serialize data
            serializer_context = {
                'inquiry': inquiry,
                'request': request
            }

            if page is not None:
                serializer = DynamicModelSerializer(
                    page,
                    many=True,
                    context=serializer_context
                )

                # Log execution
                execution_time = int((time.time() - start_time) * 1000)
                self.log_execution(
                    inquiry=inquiry,
                    user=request.user,
                    filters=filters,
                    search=search,
                    sort=sort,
                    result_count=paginator.page.paginator.count,
                    page_size=len(page),
                    page_number=request.GET.get('page', 1),
                    execution_time=execution_time,
                    query_count=builder.query_count,
                    request=request
                )

                response_data = paginator.get_paginated_response(serializer.data).data
                response_data['aggregations'] = aggregations
                response_data['execution_time_ms'] = execution_time
                response_data['inquiry'] = {
                    'name': inquiry.name,
                    'display_name': inquiry.display_name,
                    'code': inquiry.code
                }

                return Response(response_data)

            # No pagination
            serializer = DynamicModelSerializer(
                queryset,
                many=True,
                context=serializer_context
            )

            execution_time = int((time.time() - start_time) * 1000)
            self.log_execution(
                inquiry=inquiry,
                user=request.user,
                filters=filters,
                search=search,
                sort=sort,
                result_count=queryset.count(),
                execution_time=execution_time,
                query_count=builder.query_count,
                request=request
            )

            return Response({
                'results': serializer.data,
                'count': queryset.count(),
                'aggregations': aggregations,
                'execution_time_ms': execution_time,
                'inquiry': {
                    'name': inquiry.name,
                    'display_name': inquiry.display_name,
                    'code': inquiry.code
                }
            })

        except InquiryConfiguration.DoesNotExist:
            return Response(
                {'error': 'Inquiry not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            # Log failed execution
            if 'inquiry' in locals():
                self.log_execution(
                    inquiry=inquiry,
                    user=request.user,
                    filters=filters if 'filters' in locals() else {},
                    success=False,
                    error_message=str(e),
                    request=request
                )
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'], url_path='schema/(?P<code>[^/.]+)')
    def schema(self, request, code=None):
        """Get inquiry schema for frontend rendering"""
        try:
            inquiry = InquiryConfiguration.objects.get(code=code)

            if not self.has_permission(request.user, inquiry):
                return Response(
                    {'error': 'Permission denied'},
                    status=status.HTTP_403_FORBIDDEN
                )

            schema = {
                'inquiry': InquiryConfigurationSerializer(inquiry).data,
                'fields': [],
                'filters': [],
                'sorts': [],
                'relations': [],
                'permissions': self.get_user_permissions(request.user, inquiry)
            }

            # Add field schemas
            for field in inquiry.fields.filter(is_visible=True).order_by('order'):
                field_schema = {
                    'field_path': field.field_path,
                    'display_name': field.display_name,
                    'field_type': field.field_type,
                    'is_sortable': field.is_sortable,
                    'is_searchable': field.is_searchable,
                    'is_filterable': field.is_filterable,
                    'is_primary': field.is_primary,
                    'width': field.width,
                    'alignment': field.alignment,
                    'format_template': field.format_template,
                }
                schema['fields'].append(field_schema)

            # Add filter schemas
            for filter_config in inquiry.filters.filter(is_visible=True).order_by('order'):
                filter_schema = {
                    'code': filter_config.code,
                    'name': filter_config.name,
                    'field_path': filter_config.field_path,
                    'operator': filter_config.operator,
                    'filter_type': filter_config.filter_type,
                    'default_value': filter_config.default_value,
                    'is_required': filter_config.is_required,
                    'is_advanced': filter_config.is_advanced,
                    'placeholder': filter_config.placeholder,
                    'help_text': filter_config.help_text,
                    'validation_rules': filter_config.validation_rules,
                }

                # Add choices for select filters
                if filter_config.filter_type in ['select', 'multiselect', 'radio']:
                    if filter_config.choices_json:
                        filter_schema['choices'] = filter_config.choices_json
                    elif filter_config.lookup_category:
                        # Get choices from lookup
                        from lookup.models import Lookup
                        choices = Lookup.objects.filter(
                            parent_lookup__name=filter_config.lookup_category,
                            active_ind=True
                        ).values('id', 'name', 'code')
                        filter_schema['choices'] = list(choices)

                schema['filters'].append(filter_schema)

            # Add sort options
            for sort in inquiry.sorts.all():
                schema['sorts'].append({
                    'field_path': sort.field_path,
                    'direction': sort.direction,
                })

            return Response(schema)

        except InquiryConfiguration.DoesNotExist:
            return Response(
                {'error': 'Inquiry not found'},
                status=status.HTTP_404_NOT_FOUND
            )

    def has_permission(self, user, inquiry):
        """Check if user has permission to execute inquiry"""
        if inquiry.is_public:
            return True

        if user.is_superuser:
            return True

        user_groups = user.groups.all()
        return inquiry.allowed_groups.filter(id__in=user_groups).exists()

    def get_user_permissions(self, user, inquiry):
        """Get user's permissions for the inquiry"""
        if user.is_superuser:
            return {
                'can_view': True,
                'can_export': True,
                'can_view_all': True,
                'export_formats': ['csv', 'excel', 'json'],
            }

        user_groups = user.groups.all()
        permission = inquiry.permissions.filter(group__in=user_groups).first()

        if permission:
            return {
                'can_view': permission.can_view,
                'can_export': permission.can_export,
                'can_view_all': permission.can_view_all,
                'export_formats': permission.allowed_export_formats or ['csv'],
                'max_export_rows': permission.max_export_rows,
            }

        return {
            'can_view': inquiry.is_public,
            'can_export': False,
            'can_view_all': False,
            'export_formats': [],
        }

    def log_execution(self, inquiry, user, filters=None, search=None,
                      sort=None, result_count=0, page_size=None,
                      page_number=None, execution_time=0, query_count=0,
                      export_format=None, success=True, error_message='',
                      request=None):
        """Log inquiry execution"""
        InquiryExecution.objects.create(
            inquiry=inquiry,
            user=user,
            filters_applied=filters or {},
            search_query=search or '',
            sort_applied=sort or [],
            result_count=result_count,
            page_size=page_size,
            page_number=page_number,
            execution_time_ms=execution_time,
            query_count=query_count,
            export_format=export_format or '',
            success=success,
            error_message=error_message,
            ip_address=self.get_client_ip(request) if request else None,
            user_agent=request.META.get('HTTP_USER_AGENT', '')[:500] if request else ''
        )

    def get_client_ip(self, request):
        """Get client IP address"""
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0]
        else:
            ip = request.META.get('REMOTE_ADDR')
        return ip

    def export_data(self, queryset, inquiry, export_format, user):
        """Export data in requested format"""
        # Check export permission
        user_groups = user.groups.all()
        permission = inquiry.permissions.filter(group__in=user_groups).first()

        if permission:
            if not permission.can_export:
                return Response(
                    {'error': 'Export not allowed'},
                    status=status.HTTP_403_FORBIDDEN
                )
            if export_format not in permission.allowed_export_formats:
                return Response(
                    {'error': f'Export format {export_format} not allowed'},
                    status=status.HTTP_403_FORBIDDEN
                )
            if permission.max_export_rows and queryset.count() > permission.max_export_rows:
                return Response(
                    {'error': f'Export exceeds maximum rows ({permission.max_export_rows})'},
                    status=status.HTTP_403_FORBIDDEN
                )

        # Serialize data
        serializer = DynamicModelSerializer(
            queryset,
            many=True,
            context={'inquiry': inquiry, 'request': None}
        )
        data = serializer.data

        # Get visible fields for export
        export_fields = inquiry.fields.filter(is_visible=True).order_by('order')

        if export_format == 'csv':
            return self.export_csv(data, export_fields, inquiry)
        elif export_format == 'excel':
            return self.export_excel(data, export_fields, inquiry)
        elif export_format == 'json':
            return self.export_json(data, inquiry)
        else:
            return Response(
                {'error': 'Invalid export format'},
                status=status.HTTP_400_BAD_REQUEST
            )

    def export_csv(self, data, fields, inquiry):
        """Export as CSV"""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{inquiry.code}_export.csv"'

        writer = csv.writer(response)

        # Write headers
        headers = [field.display_name for field in fields]
        writer.writerow(headers)

        # Write data
        for row in data:
            row_data = []
            for field in fields:
                value = row.get(field.field_path, '')
                if isinstance(value, (dict, list)):
                    value = json.dumps(value)
                row_data.append(value)
            writer.writerow(row_data)

        return response

    def export_excel(self, data, fields, inquiry):
        """Export as Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = inquiry.display_name[:31]  # Excel sheet name limit

        # Write headers
        headers = [field.display_name for field in fields]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # Write data
        for row_idx, row in enumerate(data, 2):
            for col_idx, field in enumerate(fields, 1):
                value = row.get(field.field_path, '')
                if isinstance(value, (dict, list)):
                    value = json.dumps(value)
                ws.cell(row=row_idx, column=col_idx, value=value)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{inquiry.code}_export.xlsx"'
        wb.save(response)

        return response

    def export_json(self, data, inquiry):
        """Export as JSON"""
        response = HttpResponse(content_type='application/json')
        response['Content-Disposition'] = f'attachment; filename="{inquiry.code}_export.json"'

        export_data = {
            'inquiry': {
                'name': inquiry.name,
                'code': inquiry.code,
                'exported_at': time.strftime('%Y-%m-%d %H:%M:%S')
            },
            'data': data,
            'count': len(data)
        }

        json.dump(export_data, response, indent=2)
        return response


class InquiryTemplateViewSet(viewsets.ModelViewSet):
    """CRUD for saved inquiry templates"""
    queryset = InquiryTemplate.objects.all()
    serializer_class = InquiryTemplateSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = super().get_queryset()
        # Filter by ownership
        return queryset.filter(
            created_by=self.request.user
        ) | queryset.filter(is_public=True)

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

class InquiryFieldViewSet(viewsets.ModelViewSet):
    """CRUD for inquiry fields"""
    queryset = InquiryField.objects.all()
    serializer_class = InquiryFieldSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = super().get_queryset()
        inquiry_id = self.request.query_params.get('inquiry')
        if inquiry_id:
            queryset = queryset.filter(inquiry_id=inquiry_id)
        return queryset.order_by('order')


class InquiryFilterViewSet(viewsets.ModelViewSet):
    """CRUD for inquiry filters"""
    queryset = InquiryFilter.objects.all()
    serializer_class = InquiryFilterSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = super().get_queryset()
        inquiry_id = self.request.query_params.get('inquiry')
        if inquiry_id:
            queryset = queryset.filter(inquiry_id=inquiry_id)
        return queryset.order_by('order')


class InquiryRelationViewSet(viewsets.ModelViewSet):
    """CRUD for inquiry relations"""
    queryset = InquiryRelation.objects.all()
    serializer_class = InquiryRelationSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = super().get_queryset()
        inquiry_id = self.request.query_params.get('inquiry')
        if inquiry_id:
            queryset = queryset.filter(inquiry_id=inquiry_id)
        return queryset.order_by('order')


class InquirySortViewSet(viewsets.ModelViewSet):
    """CRUD for inquiry sorts"""
    queryset = InquirySort.objects.all()
    serializer_class = InquirySortSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = super().get_queryset()
        inquiry_id = self.request.query_params.get('inquiry')
        if inquiry_id:
            queryset = queryset.filter(inquiry_id=inquiry_id)
        return queryset.order_by('order')

    @action(detail=False, methods=['post'])
    def bulk_update(self, request):
        """Bulk update sort orders"""
        inquiry_id = request.data.get('inquiry')
        sorts_data = request.data.get('sorts', [])

        if not inquiry_id:
            return Response({'error': 'inquiry field is required'}, status=status.HTTP_400_BAD_REQUEST)

        # Delete existing sorts
        InquirySort.objects.filter(inquiry_id=inquiry_id).delete()

        # Create new sorts
        sorts = []
        for idx, sort_data in enumerate(sorts_data):
            sort_obj = InquirySort.objects.create(
                inquiry_id=inquiry_id,
                field_path=sort_data['field_path'],
                direction=sort_data.get('direction', 'asc'),
                order=idx
            )
            sorts.append(sort_obj)

        serializer = InquirySortSerializer(sorts, many=True)
        return Response(serializer.data)


class InquiryPermissionViewSet(viewsets.ModelViewSet):
    """CRUD for inquiry permissions"""
    queryset = InquiryPermission.objects.all()
    serializer_class = InquiryPermissionSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = super().get_queryset()
        inquiry_id = self.request.query_params.get('inquiry')
        if inquiry_id:
            queryset = queryset.filter(inquiry_id=inquiry_id)
        return queryset

    @action(detail=False, methods=['post'])
    def bulk_update(self, request):
        """Bulk update permissions"""
        permissions_data = request.data

        if not isinstance(permissions_data, list):
            return Response({'error': 'Expected list of permissions'}, status=status.HTTP_400_BAD_REQUEST)

        updated_permissions = []
        for perm_data in permissions_data:
            perm_id = perm_data.get('id')
            if perm_id:
                # Update existing
                try:
                    permission = InquiryPermission.objects.get(id=perm_id)
                    serializer = InquiryPermissionSerializer(permission, data=perm_data, partial=True)
                    if serializer.is_valid():
                        serializer.save()
                        updated_permissions.append(serializer.data)
                except InquiryPermission.DoesNotExist:
                    pass
            else:
                # Create new
                serializer = InquiryPermissionSerializer(data=perm_data)
                if serializer.is_valid():
                    serializer.save()
                    updated_permissions.append(serializer.data)

        return Response(updated_permissions)